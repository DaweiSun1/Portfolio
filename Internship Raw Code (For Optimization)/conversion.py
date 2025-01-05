import openpyxl.styles
import pandas as pd
from tkinter import Tk, Button, filedialog, messagebox, Frame, Checkbutton, BooleanVar
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.formatting.rule import CellIsRule 
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import re
import shutil
import numpy as np
from fpdf import FPDF


def convert_and_process_file(show_prompts):
    try:
        # Prompt the user to select the raw .xls file
        if show_prompts:
             messagebox.showinfo("Select Directory", "Please select the main directory where the .xsl file is located.")
        raw_file_path = filedialog.askopenfilename(
            title="Select Raw .xls File", 
            filetypes=[("Excel files", "*.xls")]
        )


        if not raw_file_path:
            if show_prompts:
                messagebox.showwarning("Warning", "No directory selected. Exiting.")
            return
        
        if show_prompts:
            messagebox.showinfo("Select Directory", "Please select the main directory where the folders are located")
        main_directory = filedialog.askdirectory(
            title = "Select Main Directory Containing All Shop Order Folders"
        )

        if not main_directory:
            if show_prompts:
                messagebox.showwarning("Warning", "No directory selected. Exiting.")
            return
        
        # Convert .xls to .xlsx
        df = pd.read_excel(raw_file_path)
        xlsx_file_path = os.path.splitext(raw_file_path)[0] + "_converted.xlsx"
        df.to_excel(xlsx_file_path, index=False)
        
        # Load the converted .xlsx file
        df_converted = pd.read_excel(xlsx_file_path)
        df_converted.drop_duplicates()
        # Select the specified columns and add the new column
        selected_columns = ['PART_NO', 'SHOP_ORDER', 'START_DATE', 'QTY_ORDER', 'Comment from Header', 'SO Due Dt', 'WC_DESC']
        df_selected = df_converted.loc[:, selected_columns].copy()
        df_selected['Quantity Left'] = ''
        df_selected['Enter Number'] = ''
        df_selected['Enter Number'] = df_selected['Enter Number'].astype(str)
        df_selected['Finish?'] = ''
        df_selected["Completeness"] = ''
        df_selected['Defect Code'] = ''
        df_selected['Initial'] = ''
        df_selected['NEXT STOP'] = df_selected.groupby(['PART_NO', 'SHOP_ORDER'])['WC_DESC'].shift(-1)

        # Save the new dataframe to a new Excel file
        output_file_path = os.path.splitext(raw_file_path)[0] + "_processed.xlsx"
        df_selected.to_excel(output_file_path, index=False)
        os.remove(xlsx_file_path)
        
        # Information for the Rev Check List
        df_1 = pd.read_excel(raw_file_path)
        xlsx_file_path_1 = os.path.splitext(raw_file_path)[0]+"_converted_1.xlsx"
        df_1.to_excel(xlsx_file_path_1,index=False)
        df_converted_1 = pd.read_excel(xlsx_file_path_1)
        df_converted_1.drop_duplicates()
        selected_columns_1 = ['PART_NO','SHOP_ORDER']
        df_selected_1 = df_converted_1.loc[:, selected_columns_1].copy()
        df_selected_1['Program Rev'] = ''
        df_selected_1['Rev 2'] = ''
        df_selected_1['Match?'] = ''
        output_file_path_1 = os.path.splitext(raw_file_path)[0] + "Rev_Check.xlsx"
        df_selected_1.to_excel(output_file_path_1, index=False)
        os.remove(xlsx_file_path_1)
        df_1 = pd.read_excel(output_file_path_1)
        unique_shop_orders_df = df_1.drop_duplicates(subset=['SHOP_ORDER'])
        unique_shop_orders_df.to_excel(output_file_path_1,index=False)
        wb_1 = load_workbook(output_file_path_1)
        # Load the processed file to format the dates
        wb = load_workbook(output_file_path)
        ws = wb.active
        ws.title = 'ORIGINAL DATA'

        dv = DataValidation(type="list", formula1='"Finished,In_Progress"', allow_blank=True)
        dv_1 = DataValidation(type="list", formula1='"Completed,Not_Completed"', allow_blank=True)
        ws.add_data_validation(dv)
        ws.add_data_validation(dv_1)
        # ws.add_data_validation(dv_defect)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=10)
            dv.add(cell) 
            cell_1 = ws.cell(row=row, column=11)
            dv_1.add(cell_1)
        # Define the date format style
        date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
        
        # Color Mapping 
        color_mapping = {
            "ORANGE": "FF8000",
            "ROYAL BLUE": "4169e1",
            "BLACK": "000000",
            "NEON GREEN": "39FF14",
            "RED": "FF0000",
            "VIOLET": "8000FF",
            "NEON BLUE": "4666FF",
            "YELLOW": "FFFF00",
            "GREEN": "00FF00",
            "PINK": "FFC0CB",
            "MAROON": "800000",
            "GRAY": "808080"
            }
        
        # Apply the date format to the START_DATE and SO Due Dt columns
        for row in ws.iter_rows(min_row=2, min_col=ws.min_column, max_col=ws.max_column):
            start_date_cell = row[2] # START_DATE
            due_date_cell = row[5] # DUE_DATE

            if start_date_cell.value is not None:  # START_DATE
                start_date_cell.style = date_style
            if due_date_cell.value is not None:  # SO Due Dt
                due_date_cell.style = date_style

            # Apply color fill based on comments
            comment_cell = row[4]  # Assuming 'Comment from Header' is the 5th column (index 4)
            part_number_cell = row[0]  # Assuming 'PART_NUMBER' is the 1st column (index 0)
            for keyword, hex_color in color_mapping.items():
                if keyword in (comment_cell.value or '').upper():
                    part_number_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')    
                    if (hex_color == "000000"):
                        part_number_cell.font = Font(color = "FFFFFF")

        # Adjust column widths for better readability
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        for ws in wb_1.worksheets:
            # Adjust column widths for better readability
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 4)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Split the data into multiple sheets based on WC_DESC
        unique_wc_desc = df_selected['WC_DESC'].unique()

        for wc_desc in unique_wc_desc:
            sheet_name = re.sub(r'[\\/*?:"<>|]', "_", wc_desc)[:31]
            ws_new = wb.create_sheet(title = sheet_name)
            group = df_selected[df_selected['WC_DESC'] == wc_desc]
            for r in dataframe_to_rows(group, index=False, header=True):
                ws_new.append(r) 
            for column in ws_new.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 4)
                ws_new.column_dimensions[column[0].column_letter].width = adjusted_width

            for row in ws_new.iter_rows(min_row=2, min_col=ws_new.min_column, max_col=ws_new.max_column):
                start_date_cell = row[2] # START_DATE 
                due_date_cell = row[5] # DUE_DATE

                if start_date_cell.value is not None:  # START_DATE
                    start_date_cell.style = date_style
                if due_date_cell.value is not None:  # SO Due Dt
                    due_date_cell.style = date_style
                # Apply color fill based on comments
                comment_cell = row[4]  # Assuming 'Comment from Header' is the 5th column (index 4)
                part_number_cell = row[0]  # Assuming 'PART_NUMBER' is the 1st column (index 0)
                
                comment_value = str(comment_cell.value).upper() if comment_cell.value is not None else ''
                for keyword, hex_color in color_mapping.items():
                    if keyword in comment_value:
                        part_number_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')    
                        if (hex_color == "000000"):
                            part_number_cell.font = Font(color = "FFFFFF")
            # Define conditional formatting for "Complete" and "Defect"
            green_fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid") 
            red_fill = PatternFill(start_color="E60026", end_color="E60026", fill_type="solid")
            yellow_fill = PatternFill(start_color= "FFEF00", end_color= "FFEF00", fill_type="solid")
            blue_fill = PatternFill(start_color="00B9E8", end_color="00B9E8", fill_type="solid")

            complete_rule = CellIsRule(operator="equal", formula=['"Completed"'], fill=green_fill)
            not_complete_rule = CellIsRule(operator="equal", formula=['"Not_Completed"'], fill=red_fill)
            progress_rule = CellIsRule(operator="equal",formula=['"In_Progress"'], fill=yellow_fill)
            finish_rule = CellIsRule(operator="equal",formula=['"Finished"'], fill = blue_fill)

            dv = DataValidation(type="list", formula1='"Finished,In_Progress"', allow_blank=True)
            dv_1 = DataValidation(type="list", formula1='"Completed,Not_Completed"', allowBlank=True)

            ws_new.add_data_validation(dv)
            ws_new.add_data_validation(dv_1)

            for row in range(2, ws.max_row + 1):
                cell = ws_new.cell(row=row, column=10)
                cell_original = ws.cell(row=row, column=10)
                cell_1 = ws_new.cell(row=row,column = 11)
                cell_original_1 = ws.cell(row=row,column=11)
                dv.add(cell) 
                dv_1.add(cell_1)
                # Add rules 
                ws_new.conditional_formatting.add(cell_1.coordinate, complete_rule)
                ws_new.conditional_formatting.add(cell_1.coordinate, not_complete_rule)
                ws_new.conditional_formatting.add(cell.coordinate, progress_rule)
                ws_new.conditional_formatting.add(cell.coordinate, finish_rule)
                ws.conditional_formatting.add(cell_original_1.coordinate, complete_rule)
                ws.conditional_formatting.add(cell_original_1.coordinate, not_complete_rule)
                ws.conditional_formatting.add(cell_original.coordinate, progress_rule)
                ws.conditional_formatting.add(cell_original.coordinate, finish_rule)

        # Add hyperlink
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    shop_order = cell.value
                    folder_path = os.path.join(main_directory, str(shop_order))
                    if os.path.exists(folder_path):
                        week_num = os.path.basename(raw_file_path)[4:6]
                        cell.hyperlink = Hyperlink(ref=cell.coordinate, target="https://charlesind-my.sharepoint.com/my?id=%2Fpersonal%2Fdsun%5Fcharlesindustries%5Fcom%2FDocuments%2FWeek%20"+week_num+"%5Ffile%2FWeek%20" + week_num + "%2F"+str(shop_order))
                        cell.style = "Hyperlink"

        for ws in wb_1.worksheets:
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    shop_order = cell.value
                    folder_path = os.path.join(main_directory, str(shop_order))
                    if os.path.exists(folder_path):
                        week_num = os.path.basename(raw_file_path)[4:6]
                        cell.hyperlink = Hyperlink(ref=cell.coordinate, target="https://charlesind-my.sharepoint.com/my?id=%2Fpersonal%2Fdsun%5Fcharlesindustries%5Fcom%2FDocuments%2FWeek%20"+week_num+"%5Ffile%2FWeek%20" + week_num + "%2F"+str(shop_order))
                        cell.style = "Hyperlink"

        # Save the workbook with formatted dates
        wb.save(output_file_path)
        wb_1.save(output_file_path_1)
        # Load the workbook and select the worksheet
        workbook = load_workbook(output_file_path)

        for sheet_name in workbook.sheetnames:

            worksheet = workbook[sheet_name]
            len_df = pd.read_excel(output_file_path, sheet_name=sheet_name)
            # Add the 'Quantity Left' column header
            worksheet.cell(row=1, column=8, value='Quantity Left')

            # Write the formula for each row in the 'Quantity Left' column
            for row_num in range(2, len(len_df) + 2):
                worksheet.cell(row=row_num, column=8, value=f'=D{row_num} - SUMPRODUCT(--TEXTSPLIT(I{row_num}, "/"))')

            # Set the format of 'Enter Number' column to text to prevent date conversion
            for cell in worksheet['I']:
                cell.number_format = '@'
            for sheet in workbook.worksheets:
                sheet.sheet_properties.pageSetUpPr.fitToPage = True
        workbook.save(output_file_path)

        data = pd.read_excel(output_file_path, sheet_name= 'ORIGINAL DATA')
        unique_shop_orders = data['SHOP_ORDER'].unique()
        for shop_order in unique_shop_orders:
            # Filter rows for the current shop order
            shop_order_data = data[data['SHOP_ORDER'] == shop_order]
            
            for index, row in shop_order_data.iterrows():
                part_number = row['PART_NO']
                quantity_ordered = row['QTY_ORDER']
                tag_pdf = generate_tags(part_number, shop_order, quantity_ordered)
                tag_path = os.path.join(main_directory + "/" + str(shop_order),  f"TAG_{part_number}_{shop_order}.pdf")
                tag_pdf.output(tag_path)
                
        PASSWORD = "Ulock!"
        unprotected_columns = ["Quantity Left","Enter Number", "Completeness", "Defect Code", "Initial","Finish?"]
        protected_column(output_file_path,output_file_path, PASSWORD, unprotected_columns)
        messagebox.showinfo("Success", f"Processed file saved as {output_file_path}")
    
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def generate_new_folders(show_prompts):
    try:
        if show_prompts:
            messagebox.showinfo("Select Directory", "Please select the main directory wherre the folders will be created")
        main_directory_path = filedialog.askdirectory(
            title = "Please Select the Main Directory"
        )

        if not main_directory_path:
            if show_prompts:
                messagebox.showwarning("Warning", "No directory selected. Exiting.")
            return 
        if show_prompts:
            messagebox.showinfo("Select File", "Please select the raw .xls file")
        raw_file_path = filedialog.askopenfilename(
            title="Select Raw .xls File", 
            filetypes=[("Excel files", "*.xls")]
        )
    
        if not raw_file_path:
            if show_prompts:
                messagebox.showwarning("Warning","No file selected. Exiting. ")
            return
            
        raw_file_name = os.path.basename(raw_file_path)
        folder_name = "Week " + raw_file_name[4:6]
        output_folder_path = os.path.join(main_directory_path,folder_name + "_file")
        os.makedirs(output_folder_path,exist_ok= False)
        main_folder_path = os.path.join(output_folder_path, folder_name)
        os.makedirs(main_folder_path, exist_ok=False)

        # Load the spreadsheet
        df = pd.read_excel(raw_file_path)

        # Create new folders for each unique order inside of the raw spreadsheet
        unique_shop_orders = df['SHOP_ORDER'].unique()
        for shop_order in unique_shop_orders:
            folder_path = os.path.join(main_folder_path, str(shop_order))
            os.makedirs(folder_path, exist_ok=True)

        messagebox.showinfo("Success", f"New Folders Have been Generated under {main_folder_path}")
    
    except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

def on_generate_folders():
    generate_new_folders(show_prompts_var.get())

def on_convert():
    convert_and_process_file(show_prompts_var.get())

def create_windows(root, width_percentage, height_percentage):
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    width = int(screen_width * width_percentage*0.4)
    height =  int(screen_height * height_percentage*0.4)
        # Calculate the position to center the window
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)

    # Set the position of the window
    root.geometry(f'{width}x{height}+{x}+{y}')

def protected_column(input_path, output_path, password, unprotected_columns):
    wb = load_workbook(input_path)

    # Loop through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == 'LABEL PRINTING':
            continue
        ws = wb[sheet_name]

        # Cretate a set of columns to be unprotected
        unprotected_cols = set()
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value in unprotected_columns:
                unprotected_cols.add(col[0].column)

        # Protect each cell based on the condition
        for row in ws.iter_rows():
            for cell in row:
              if cell.column not in unprotected_cols:
                  cell.protection = openpyxl.styles.Protection(locked=True)
              else:
                    cell.protection = openpyxl.styles.Protection(locked=False)
        # Protect the sheet with password
        ws.protection.sheet = True
        ws.protection.set_password(password)
    
    wb.save(output_path)

def get_formula_for_quantity_left(row_num):
    return f'=D{row_num} - SUMPRODUCT(--TEXTSPLIT(I{row_num}, "/"))'

def generate_tags(part_number, shop_order, quantity_ordered):
    pdf = FPDF('P', 'mm', 'A4')
    pdf.add_page()
    
    # Title
    pdf.set_font("Arial", size=24, style='B')
    pdf.cell(0, 40, txt="SHOP ORDER ID TAG", ln=True, align='C')
    
    # Part Number
    pdf.set_font("Arial", size=32)
    pdf.cell(0, 40, txt=f"PART #: {part_number}", ln=True, align='C')
    
    # Shop Order Number
    pdf.set_font("Arial", size=32)
    pdf.cell(0, 40, txt=f"S/O #: {shop_order}", ln=True, align='C')
    
    # Quantity
    pdf.set_font("Arial", size=32)
    pdf.cell(0, 40, txt=f"QTY: {quantity_ordered}", ln=True, align='C')
    
    # Identification
    pdf.set_font("Arial", size=24)
    pdf.cell(0, 40, txt= "Prior to Assembly Use only", ln=True, align='C')
    
    return pdf

def match_color(row):
    if pd.isna(row['Program Rev']) or pd.isna(row['Rev 2']):
        return ''
    elif row['Program Rev'] == row['Rev 2']
        return 'green'
    else:
        return 'red'
    
# Set up the GUI
root = Tk()
root.title("Excel File Converter")
width_percent = 0.5
height_percent = 0.5
create_windows(root,width_percent, height_percent)

frame = Frame(root, padx=20, pady=20)
frame.pack(padx=10, pady=10)

show_prompts_var = BooleanVar(value=True)
Checkbutton(frame, text = "Show Prompts", variable=show_prompts_var).pack()

button1 = Button(frame, text="Generate New Empty Folders Based on the Shop Order", command=on_generate_folders)
button1.pack()

button2 = Button(frame, text="Convert and Process Excel File", command=on_convert)
button2.pack()

# Run the GUI loop
root.mainloop()

