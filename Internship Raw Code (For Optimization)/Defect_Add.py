import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

import pandas as pd
from tkinter import Tk, Button, filedialog, messagebox, Frame, Checkbutton, BooleanVar
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.formatting.rule import CellIsRule 
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import re
import shutil
import numpy as np
import datetime

class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Selector")

        # Initialize the filename and data
        self.filename = ""
        self.temp_filename = ""
        self.filtered_data = pd.DataFrame()  # To hold data from all sheets that match the criteria
        self.selected_directory = ""  # To store the selected directory for Week XX_file
        self.added_part_nos = set()  # Track added PART_NO to avoid duplicates

        # Treeview for displaying Excel data
        self.tree = ttk.Treeview(root)
        self.tree.pack(expand=True, fill=tk.BOTH)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        # Define tag for blue background
        self.tree.tag_configure('blue', background='lightblue')

        # Button to select a file
        self.select_file_button = tk.Button(root, text="Select File", command=self.select_file)
        self.select_file_button.pack(pady=5)

        # Button to select a directory for Week XX_file
        self.select_directory_button = tk.Button(root, text="Select Week Directory", command=self.select_directory)
        self.select_directory_button.pack(pady=5)

        # Button to add selected line
        self.add_button = tk.Button(root, text="Add", command=self.add_selected_row, state=tk.DISABLED)
        self.add_button.pack(pady=5)

    def select_file(self):
        self.filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.filename:
            try:
                self.create_temp_copy()
                self.load_and_filter_sheets()
                self.display_data()
            except PermissionError:
                messagebox.showerror("Permission Error", f"Failed to load file: Permission denied for file {self.filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load file: {e}")

    def select_directory(self):
        directory = filedialog.askdirectory(title="Select Week Directory")
        if directory:
            self.selected_directory = directory
            week_number = os.path.basename(self.selected_directory)
            week_num = week_number[6:8]
            if not week_number.startswith("Week ") or not week_number.endswith("_file"):
                messagebox.showerror("Invalid Directory", "Please select a directory named as 'Week XX_file'.")
                self.selected_directory = ""
            else:
                self.load_and_filter_sheets()  # Load and filter sheets after directory selection

    def create_temp_copy(self):
        # Create a temporary directory
        temp_dir = tempfile.mkdtemp()
        # Create a temporary copy of the file
        self.temp_filename = os.path.join(temp_dir, os.path.basename(self.filename))
        shutil.copy2(self.filename, self.temp_filename)

    def load_and_filter_sheets(self):
        try:
            # Load the workbook using openpyxl to access values
            workbook = load_workbook(self.temp_filename, data_only=True)
            filtered_rows = []
            seen_shop_orders = set()  # Track seen SHOP_ORDERs
            self.added_part_nos = self.get_existing_part_nos()  # Retrieve existing PART_NO from directories

            # Iterate over all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Identify column indices for 'Finish?', 'Completeness', 'SHOP_ORDER', and 'PART_NO'
                headers = [cell.value for cell in sheet[1]]
                finish_index = headers.index('Finish?')
                completeness_index = headers.index('Completeness')
                shop_order_index = headers.index('SHOP_ORDER')
                part_no_index = headers.index('PART_NO')

                # Iterate over all rows in the sheet
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming row 1 is the header
                    shop_order = row[shop_order_index]
                    part_no = row[part_no_index]
                    if (
                        shop_order not in seen_shop_orders and
                        row[finish_index] == "Finished" and
                        row[completeness_index] == "Not_Completed"
                    ):
                        filtered_rows.append((list(row) + [sheet_name], part_no))
                        seen_shop_orders.add(shop_order)  # Mark this SHOP_ORDER as seen

            # Create a DataFrame from the filtered rows
            columns = headers + ['Worksheet']  # Including the sheet name as a column
            self.filtered_data = pd.DataFrame([row for row, _ in filtered_rows], columns=columns)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load and filter sheets: {e}")

    def get_existing_part_nos(self):
        existing_part_nos = set()
        if self.selected_directory:
            week_number = os.path.basename(self.selected_directory).replace('_file', '')
            defected_order_path = os.path.join(self.selected_directory, f"{week_number} Defected Order")
            if os.path.exists(defected_order_path):
                for folder_name in os.listdir(defected_order_path):
                    if folder_name.endswith("_defect"):
                        part_no = folder_name.replace("_defect", "")
                        existing_part_nos.add(part_no)
        return existing_part_nos

    def display_data(self):
        # Clear the current data in the tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Set up the columns and headers
        self.tree["columns"] = list(self.filtered_data.columns)
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)

        # Insert the data and apply blue tag to existing items
        for index, row in self.filtered_data.iterrows():
            part_no = row['PART_NO']
            if part_no in self.added_part_nos:
                self.tree.insert("", "end", values=list(row), tags=('blue',))
            else:
                self.tree.insert("", "end", values=list(row))

        if self.filtered_data.empty:
            messagebox.showinfo("Information", "No rows found with the specified criteria.")
        else:
            self.add_button.config(state=tk.NORMAL)  # Enable the add button if there are rows
            messagebox.showinfo("Success", "Data loaded and displayed successfully.")

    def on_tree_select(self, event):
        selected_item = self.tree.selection()
        if selected_item:
            self.selected_row = self.tree.item(selected_item)["values"]

    def add_selected_row(self):
        if self.selected_directory and self.selected_row:
            week_number = os.path.basename(self.selected_directory).replace('_file', '')
            week_num = week_number[5:7]
            defected_order_path = os.path.join(self.selected_directory, f"{week_number} Defected Order")
            part_no = self.selected_row[0]  # Assuming PART_NO is the first column
            shop_order = self.selected_row[1]  # Assuming SHOP_ORDER is the second column

            # Create the directory if it doesn't exist
            if not os.path.exists(defected_order_path):
                os.makedirs(defected_order_path)

            # Create a folder named as PART_NO_defect
            part_no_folder = os.path.join(defected_order_path, f"{part_no}_defect")
            if not os.path.exists(part_no_folder):
                os.makedirs(part_no_folder)
                messagebox.showinfo("Success", f"Folder '{part_no_folder}' created successfully.")
            else:
                messagebox.showinfo("Information", f"Folder '{part_no_folder}' already exists.")

            # Create an Excel file within the folder
            self.create_excel_file(part_no_folder, part_no, shop_order, week_num)

            # Remove the added row from the tree
            for item in self.tree.get_children():
                if self.tree.item(item)["values"] == self.selected_row:
                    self.tree.item(item, tags=('blue',))  # Apply the 'marked' tag
                    break

            # Mark this part number as added
            self.added_part_nos.add(part_no)

    def create_excel_file(self, directory, part_no, shop_order, week_num):
        new_file_path = os.path.join(directory, f"{part_no}_defect.xlsx")

        # Load the original workbook
        workbook = load_workbook(self.temp_filename, data_only=True)

        # Create a new workbook
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)  # Remove the default sheet

        # Iterate over each sheet in the original workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []

            # Get headers
            headers = [cell.value for cell in sheet[1]]
            part_no_index = headers.index('PART_NO')
            shop_order_index = headers.index('SHOP_ORDER')
            order_quantity_index = headers.index('QTY_ORDER')
            enter_number_index = headers.index('Enter Number')
            Finish_index = headers.index('Finish?')
            Complete_index = headers.index('Completeness')
            Comment_index = headers.index('Comment from Header')

            # Extract rows that match the PART_NO and SHOP_ORDER
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[part_no_index] == part_no and row[shop_order_index] == shop_order:
                    row_data = list(row)
                    row_data[shop_order_index] = ''
                    row_data[Finish_index] = ''
                    row_data[Complete_index] = ''
                    row_data[Comment_index] = ''
                    
                    # Handle Enter Number
                    enter_number_value = self.selected_row[enter_number_index]
                    if isinstance(enter_number_value, str):
                        try:
                        # Split by '/' and sum the parts
                            enter_number_sum = sum(map(int, enter_number_value.split('/')))
                        except ValueError:
                            enter_number_sum = 0  # Handle invalid entries gracefully
                    else:
                        enter_number_sum = int(enter_number_value) if enter_number_value else 0

                    # Calculate the new QTY_ORDER
                    original_qty = row[order_quantity_index]
                    if isinstance(original_qty, (int, float)):
                        row_data[order_quantity_index] = original_qty - enter_number_sum
                    else:
                        row_data[order_quantity_index] = original_qty  # Keep original if not numeric
                    row_data[enter_number_index] = 0

                    data.append(row_data)

            # Create a new sheet in the new workbook
            new_sheet = new_workbook.create_sheet(title=sheet_name)

            # Write headers
            new_sheet.append(headers)

            # Write data
            for row in data:
                new_sheet.append(row)

            if (new_sheet.max_row <= 1):
                new_workbook.remove(new_sheet)

        adjust_column_widths(new_workbook)
        # Save the new workbook
        formatting(new_workbook)      
        for sheet_name in new_workbook.sheetnames:
            sheet = new_workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            part_no_index = headers.index('PART_NO') + 1
            shop_order = self.tree.item(self.tree.selection()) ["values"][1]
            for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=part_no_index,max_col=part_no_index):
                row[0].hyperlink = Hyperlink(ref=row[0].coordinate, target="https://charlesind-my.sharepoint.com/my?id=%2Fpersonal%2Fdsun%5Fcharlesindustries%5Fcom%2FDocuments%2FWeek%20"+week_num+"%5Ffile%2FWeek%20" + week_num + "%2F"+str(shop_order))
                row[0].style = "Hyperlink"

        new_workbook.save(new_file_path)
        messagebox.showinfo("Success", f"Excel file '{new_file_path}' created with relevant data.")
   
    def extract_part_no_hyperlink(self, part_no):
        workbook = load_workbook(self.filename, data_only=False)  # Load with formulas and hyperlinks
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            part_no_index = headers.index('PART_NO')

            for row in sheet.iter_rows(min_row=2, values_only=False):  # Keep formatting and hyperlinks
                if row[part_no_index].value == part_no:
                    # Return the hyperlink for the PART_NO cell if available
                    return row[part_no_index].hyperlink.target if row[part_no_index].hyperlink else None
        return None

def adjust_column_widths(workbook: Workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 4
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

def formatting(workbook: Workbook):
    # Define conditional formatting for "Complete" and "Defect"
    green_fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid") 
    red_fill = PatternFill(start_color="E60026", end_color="E60026", fill_type="solid")
    yellow_fill = PatternFill(start_color= "FFEF00", end_color= "FFEF00", fill_type="solid")
    blue_fill = PatternFill(start_color="00B9E8", end_color="00B9E8", fill_type="solid")

    complete_rule = CellIsRule(operator="equal", formula=['"Completed"'], fill=green_fill)
    not_complete_rule = CellIsRule(operator="equal", formula=['"Not_Completed"'], fill=red_fill)
    progress_rule = CellIsRule(operator="equal",formula=['"In_Progress"'], fill=yellow_fill)
    finish_rule = CellIsRule(operator="equal",formula=['"Finished"'], fill = blue_fill)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        dv = DataValidation(type="list", formula1='"Finished,In_Progress"', allow_blank=True)
        dv_1 = DataValidation(type="list", formula1='"Completed,Not_Completed"', allow_blank=True)
        sheet.add_data_validation(dv)
        sheet.add_data_validation(dv_1)
        # Define the date format style
        date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
        
        if "datetime" not in workbook.named_styles:
            workbook.add_named_style(date_style)

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=10)
            dv.add(cell) 
            cell_1 = sheet.cell(row=row, column=11)
            dv_1.add(cell_1)
            sheet.conditional_formatting.add(cell.coordinate, progress_rule)
            sheet.conditional_formatting.add(cell.coordinate, finish_rule)
            sheet.conditional_formatting.add(cell_1.coordinate,complete_rule)
            sheet.conditional_formatting.add(cell_1.coordinate,not_complete_rule)
            
            start_date_cell = sheet.cell(row=row, column = 3) # START_DATE
            due_date_cell = sheet.cell(row=row, column=6) # DUE_DATE
            part_no_cell = sheet.cell(row=row,column=1) # PART_NP
            
            # Ensure the cell values are recognized as dates
            if isinstance(start_date_cell.value, str):
                start_date_cell.value = datetime.datetime.strptime(start_date_cell.value, '%m/%d/%Y')

            if isinstance(due_date_cell.value, str):
                due_date_cell.value = datetime.datetime.strptime(due_date_cell.value, '%m/%d/%Y')

            start_date_cell.style = date_style
            due_date_cell.style = date_style
            
            quantity_left_cell = sheet.cell(row=row, column=8)
            quantity_left_cell.value = f'=D{row} - SUMPRODUCT(--TEXTSPLIT(I{row}, "/"))'


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()
